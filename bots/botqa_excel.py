import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment,Font

# URLs de búsqueda
urls = [
    "https://www.elempleo.com/co/ofertas-empleo/trabajo-analista-de-pruebas-de-software---testing",
    "https://www.elempleo.com/co/ofertas-empleo/trabajo-analista-de-qa-de-software",
    "https://www.elempleo.com/co/ofertas-empleo/trabajo-remoto"
]

# Posibles selectores de ofertas
posibles_selectores = ["div.result-item", "div.card-body", "article", "div.info"]

# Función para buscar ofertas
def buscar_ofertas_elempleo(urls):
    ofertas = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        for url in urls:
            page.goto(url, timeout=60000)
            jobs = []

            for selector in posibles_selectores:
                try:
                    page.wait_for_selector(selector, timeout=10000)
                    jobs = page.query_selector_all(selector)
                    if jobs:
                        print(f"✅ Usando selector: {selector} en {url}")
                        break
                except:
                    continue

            if not jobs:
                print(f"⚠️ No se encontraron ofertas en {url}")
                continue

            for job in jobs:
                cargo = job.query_selector("a").inner_text().strip() if job.query_selector("a") else "N/A"
                empresa = job.query_selector("span").inner_text().strip() if job.query_selector("span") else "N/A"
                enlace = job.query_selector("a").get_attribute("href") if job.query_selector("a") else url
                if enlace and not enlace.startswith("http"):
                    enlace = "https://www.elempleo.com" + enlace
                fecha = job.query_selector("span.info-publish-date").inner_text().strip() if job.query_selector("span.info-publish-date") else "N/A"

                ofertas.append({
                    "Cargo": cargo,
                    "Empresa": empresa,
                    "Enlace": enlace,
                    "Fecha": fecha,
                    "Fuente": url
                })

        browser.close()
    return ofertas

# Función para generar Excel con estilo
def generar_excel(ofertas, nombre_archivo="BotQA-Excel.xlsx"):
    if ofertas:
        # Crear el Excel con pandas
        df = pd.DataFrame(ofertas)
        df.to_excel(nombre_archivo, index=False)

        # Abrir el archivo con openpyxl para aplicar estilos
        wb = load_workbook(nombre_archivo)
        ws = wb.active

        # Definir el color azul claro
        azul_claro = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        encabezado_font = Font(bold=True, color="FFFFFF")  # negrita y blanco

        # Aplicar color y centrar los encabezados
        for cell in ws[1]:
            cell.fill = azul_claro
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = encabezado_font

        # Ajustar el ancho de las columnas según el contenido
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # letra de la columna
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # margen extra
            ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(nombre_archivo)
        print(f"✅ Archivo Excel generado con encabezados en azul claro, centrados y columnas ajustadas: {nombre_archivo}")
    else:
        print("⚠️ No se encontraron ofertas en las páginas dadas.")

# Punto de entrada
if __name__ == "__main__":
    print("🔍 Buscando ofertas QA en elempleo.com...")
    ofertas = buscar_ofertas_elempleo(urls)
    generar_excel(ofertas)
